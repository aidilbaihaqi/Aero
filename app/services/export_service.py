"""
export_service.py — Export data dari database ke XLSX format segitiga.
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.flight import FlightFare, ScrapeRun


EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")


def _ensure_exports_dir():
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def _sanitize_sheet_name(name: str) -> str:
    """Max 31 chars, no invalid characters."""
    for c in ['\\', '/', '?', '*', '[', ']', ':']:
        name = name.replace(c, '-')
    return name[:31]


def export_triangle_xlsx(
    db: Session,
    origin: str,
    destination: str,
    start_date: date | None = None,
    end_date: date | None = None,
    scrape_date_filter: date | None = None,
) -> str:
    """
    Query data dari DB dan export ke XLSX format segitiga.

    Baris = scrape_date, Kolom = travel_date, Value = harga termurah.
    1 sheet per airline.

    Returns:
        str: absolute path ke file XLSX yang dihasilkan.
    """
    _ensure_exports_dir()
    route = f"{origin}-{destination}"

    # Query data — JOIN ScrapeRun untuk scrape_date
    query = db.query(FlightFare, ScrapeRun.scrape_date).join(
        ScrapeRun, FlightFare.run_id == ScrapeRun.run_id
    ).filter(
        FlightFare.route == route,
        FlightFare.status_scrape == "SUCCESS",
    )

    if start_date:
        query = query.filter(FlightFare.travel_date >= start_date)
    if end_date:
        query = query.filter(FlightFare.travel_date <= end_date)
    if scrape_date_filter:
        query = query.filter(ScrapeRun.scrape_date == scrape_date_filter)

    rows = query.order_by(ScrapeRun.scrape_date, FlightFare.travel_date).all()

    if not rows:
        return ""

    # Group by airline -> scrape_date -> travel_date -> cheapest price
    data: dict[str, dict[date, dict[date, float]]] = {}
    all_travel_dates_set: set[date] = set()
    for fare, scrape_dt in rows:
        airline = fare.airline
        if airline not in data:
            data[airline] = {}
        if scrape_dt not in data[airline]:
            data[airline][scrape_dt] = {}
        td = fare.travel_date
        all_travel_dates_set.add(td)
        price = float(fare.basic_fare)
        if td not in data[airline][scrape_dt] or price < data[airline][scrape_dt][td]:
            data[airline][scrape_dt][td] = price

    all_travel_dates = sorted(all_travel_dates_set)

    # Build XLSX
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    date_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for airline in sorted(data.keys()):
        sheet_name = _sanitize_sheet_name(f"{route} {airline}")
        ws = wb.create_sheet(sheet_name)

        # Header row
        cell = ws.cell(row=1, column=1, value="Scrape Date \\ Flight Date")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 18

        for col_idx, td in enumerate(all_travel_dates, start=2):
            cell = ws.cell(row=1, column=col_idx, value=td.strftime("%Y-%m-%d"))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # Data rows (per scrape_date)
        for row_idx, sd in enumerate(sorted(data[airline].keys()), start=2):
            cell = ws.cell(row=row_idx, column=1, value=sd.strftime("%Y-%m-%d"))
            cell.font = Font(bold=True, size=10)
            cell.fill = date_fill
            cell.border = border

            for col_idx, td in enumerate(all_travel_dates, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                price = data[airline][sd].get(td)
                if price:
                    cell.value = price
                    cell.number_format = '#,##0'
                else:
                    cell.value = "-"
                cell.border = border
                cell.alignment = Alignment(horizontal="right")

    # Save
    start_str = start_date.strftime("%Y-%m-%d") if start_date else (min(all_travel_dates).strftime("%Y-%m-%d") if all_travel_dates else "ALL")
    end_str = end_date.strftime("%Y-%m-%d") if end_date else (max(all_travel_dates).strftime("%Y-%m-%d") if all_travel_dates else "ALL")

    filename = f"aero_{route}_{start_str}_{end_str}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    return filepath
